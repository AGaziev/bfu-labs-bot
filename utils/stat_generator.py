from io import BytesIO

import openpyxl
from openpyxl.cell import Cell
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter


class StatsGenerator:
    lightRed = '00FF8080'
    unregisteredFill = PatternFill(patternType='solid', fgColor=lightRed)

    lightGreen = '0099CC00'
    lightBlue = '0099CCFF'
    lab_status_fill = {
        1: PatternFill(patternType='solid', fgColor=lightGreen),
        2: PatternFill(patternType='solid', fgColor=lightRed),
        3: PatternFill(patternType='solid', fgColor=lightBlue)
    }

    font_bold = Font(bold=True)

    @classmethod
    def generate_stats(cls, group_name: str, lab_stats: dict[str: [bool, [[int, str, int]]]], lab_num: int):
        """
        lab_stats: every lab like
            {student name: [is_registered, [[number, date, status_id]]]}
        """
        book = openpyxl.Workbook()
        book.remove(book.active)
        stats_sheet = book.create_sheet(f"{group_name}")

        stats_sheet.append(cls.get_lab_row(stats_sheet, lab_num))
        for i, (name, info) in enumerate(lab_stats.items()):
            stats_sheet.append(cls.get_student_row(stats_sheet, i, name, info[0], info[1]))

        cls.adjust_columns_size(stats_sheet)
        book.save("book.xlsx")
        return cls.convert_to_bytesio_file(book)

    @classmethod
    def get_student_row(cls, sheet, row_num: int, name: str, is_registered: bool,
                        labs_info: [[int, str, int]]):
        name_cell = Cell(sheet, column="A", row=row_num + 1, value=name)
        if not is_registered:
            name_cell.fill = cls.unregisteredFill
        name_cell.font = cls.font_bold
        yield name_cell
        for lab in labs_info:
            lab_cell = Cell(sheet, column=get_column_letter(lab[0] + 1), row=row_num + 1, value=lab[1])
            print(lab)
            lab_cell.fill = cls.lab_status_fill[lab[2]]
            yield lab_cell

    @classmethod
    def get_lab_row(cls, sheet, lab_num):
        descr_cell = Cell(sheet, column=1, row=1, value="ФИО")
        descr_cell.font = cls.font_bold
        yield
        for i in range(1, lab_num + 1):
            lab_cell = Cell(sheet, column=get_column_letter(i + 1), row=1, value=f"LAB{i}")
            lab_cell.font = cls.font_bold
            yield lab_cell

    @classmethod
    def adjust_columns_size(cls, sheet):
        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter  # Get the column name
            for cell in col:
                try:  # Necessary to avoid error on empty cells
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.1
            sheet.column_dimensions[column].width = adjusted_width

    @classmethod
    def convert_to_bytesio_file(cls, book: openpyxl.Workbook):
        file_ = BytesIO()
        book.save(file_)
        file_.seek(0)
